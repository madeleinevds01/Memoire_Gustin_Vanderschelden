"""Loading and statistics for data/EuroSpaceCenter-Questionnaires.xlsx.

The sheet is laid out transposed: one row per questionnaire item, one column
per participant (identified by "N deg Identifiant" in row 1). It contains, in
order, the SPATIAL WALKING section (26-item emotional/somatic scale, then the
32-item Presence Questionnaire, Witmer & Singer 1998) and then a ROTOR
section (a different, independent experiment described in the protocol as
out of scope for this thesis).

The parser below does not hardcode row numbers. It walks the sheet and
detects section boundaries (a label row with no data, e.g. "SPATIAL WALKING")
and, within a section, detects where a new questionnaire starts (item
numbering resets to "1."). That makes the two block sizes (26 and 32) a
cross-check derived from the data rather than an assumption baked into the
code.

Caveat carried into every reliability/score computed here: the Presence
Questionnaire mixes positively and negatively worded items, and this project
does not currently have a confirmed reverse-scoring key for it. Scale scores
and Cronbach's alpha below are computed on the RAW item values. If a
reverse-coding list is available (from Witmer & Singer 1998 or the specific
version used in the protocol), it should be applied before scoring -- see
`REVERSE_SCORED_PRESENCE_ITEMS` below.
"""
from __future__ import annotations

import re

import numpy as np
import openpyxl
import pandas as pd

from stats_utils import describe_series, cronbach_alpha, item_total_correlations, bootstrap_ci

ITEM_PATTERN = re.compile(r"^\s*(\d+)\.\s*(.+?)\s*$")

# Fill this in with the confirmed reverse-scored item labels (exact text as
# it appears in the sheet) before relying on the Presence Questionnaire total
# for anything beyond exploratory, item-level description. Left empty on
# purpose: guessing the wrong items would silently bias every score below.
REVERSE_SCORED_PRESENCE_ITEMS: list[str] = []


def _clean_label(value) -> str | None:
    if value is None:
        return None
    return str(value).replace("\xa0", " ").strip()


def parse_workbook(path: str, sheet_name: str = "Feuil1") -> dict:
    """Parse the workbook into {section_name: [block_dataframe, ...]}.

    Each block_dataframe has one row per participant (index = participant
    id, int) and one column per item (column name = the item label,
    including its original numbering, e.g. "1. Isolation and loneliness").
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    max_col = ws.max_column
    max_row = ws.max_row

    participant_ids = [ws.cell(row=1, column=c).value for c in range(2, max_col + 1)]

    sections: dict[str, list[list[tuple[int, str, int]]]] = {}
    current_section = None
    current_block: list[tuple[int, str, int]] = []

    def flush_block():
        nonlocal current_block
        if current_section is not None and current_block:
            sections[current_section].append(current_block)
        current_block = []

    for r in range(2, max_row + 1):
        label = _clean_label(ws.cell(row=r, column=1).value)

        if label is None:
            continue

        match = ITEM_PATTERN.match(label)
        if match is None:
            # A row whose label is not "N. <text>" is a section marker
            # (e.g. "SPATIAL WALKING", "ROTOR"), not a questionnaire item.
            # Section markers are detected purely from the label, not from
            # whether the row happens to carry stray data in some column
            # (row 61, "ROTOR", has a stray "2" in one participant column
            # that must not be mistaken for Presence Questionnaire data).
            flush_block()
            current_section = label
            sections.setdefault(current_section, [])
            continue

        item_number = int(match.group(1))
        if item_number == 1 and current_block:
            flush_block()
        current_block.append((item_number, label, r))

    flush_block()

    result = {}
    for section_name, blocks in sections.items():
        block_dfs = []
        for block in blocks:
            columns = [label for _, label, _ in block]
            rows = [r for _, _, r in block]
            data = {}
            for col_label, r in zip(columns, rows):
                values = [ws.cell(row=r, column=c).value for c in range(2, max_col + 1)]
                data[col_label] = values
            df = pd.DataFrame(data, index=pd.Index(participant_ids, name="participant_id"))
            df = df.apply(pd.to_numeric, errors="coerce")
            block_dfs.append(df)
        result[section_name] = block_dfs
    return result


def load_spatial_walking_questionnaires(path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return (emotional_df, presence_df) for the SPATIAL WALKING section only.

    The ROTOR section (a separate experiment, see module docstring) is
    parsed too but intentionally not returned here.
    """
    sections = parse_workbook(path)
    spatial_key = next((k for k in sections if "SPATIAL" in k.upper()), None)
    if spatial_key is None:
        raise ValueError(f"Could not find the SPATIAL WALKING section; sections found: {list(sections)}")
    blocks = sections[spatial_key]
    by_size = {df.shape[1]: df for df in blocks}
    if 26 not in by_size or 32 not in by_size:
        sizes = [df.shape[1] for df in blocks]
        raise ValueError(
            f"Expected a 26-item block (emotional scale) and a 32-item block "
            f"(Presence Questionnaire) in the SPATIAL WALKING section, found "
            f"blocks of size {sizes} instead. The sheet layout may have changed."
        )
    return by_size[26], by_size[32]


def item_stats_table(item_df: pd.DataFrame) -> pd.DataFrame:
    """Per-item descriptive statistics (one row per questionnaire item)."""
    rows = {col: describe_series(item_df[col]) for col in item_df.columns}
    table = pd.DataFrame(rows).T
    table.index.name = "item"
    return table.sort_values("mean", ascending=False)


def scale_scores(item_df: pd.DataFrame, min_items: int = 1) -> pd.DataFrame:
    """Per-participant scale score = mean of the items they answered.

    Using the mean of available items (rather than the sum) keeps the score
    on the original response scale regardless of how many items a given
    participant skipped. `min_items` sets how many valid answers are
    required before a score is reported at all (default 1, matching how the
    protocol's own participant counts -- 44 / 43 -- were derived: anyone
    with at least one answered item counts as a respondent).
    """
    n_items_used = item_df.notna().sum(axis=1)
    mean_score = item_df.mean(axis=1, skipna=True)
    out = pd.DataFrame({
        "n_items_used": n_items_used,
        "n_items_total": item_df.shape[1],
        "score_mean": mean_score,
    })
    out.loc[n_items_used < min_items, "score_mean"] = np.nan
    return out


def reliability_report(item_df: pd.DataFrame) -> dict:
    alpha = cronbach_alpha(item_df)
    item_total = item_total_correlations(item_df)
    return {"alpha": alpha, "item_total_correlations": item_total}


def scale_correlation(scores_a: pd.Series, scores_b: pd.Series) -> dict:
    """Pearson + Spearman correlation between two per-participant scale scores.

    Restricted to participants who have both scores (paired, listwise). A
    bootstrap CI is reported alongside Pearson's r because n here is the
    number of participants (dozens, not thousands), too small to trust a
    normal-theory CI blindly.
    """
    paired = pd.concat([scores_a.rename("a"), scores_b.rename("b")], axis=1).dropna()
    if len(paired) < 3:
        return {"n": len(paired), "pearson_r": np.nan, "pearson_p": np.nan,
                "spearman_r": np.nan, "spearman_p": np.nan,
                "pearson_ci95_low": np.nan, "pearson_ci95_high": np.nan}
    from scipy import stats as scipy_stats

    pearson_r, pearson_p = scipy_stats.pearsonr(paired["a"], paired["b"])
    spearman_r, spearman_p = scipy_stats.spearmanr(paired["a"], paired["b"])

    def _corr(a, b):
        return np.corrcoef(a, b)[0, 1]

    res = scipy_stats.bootstrap(
        (paired["a"].to_numpy(), paired["b"].to_numpy()), _corr,
        n_resamples=5000, paired=True, method="BCa", random_state=0,
    )
    return {
        "n": len(paired),
        "pearson_r": float(pearson_r), "pearson_p": float(pearson_p),
        "spearman_r": float(spearman_r), "spearman_p": float(spearman_p),
        "pearson_ci95_low": float(res.confidence_interval.low),
        "pearson_ci95_high": float(res.confidence_interval.high),
    }
